import math
import sys
import time
import openpyxl
import glob

print("2^nにおけるnを入力してください。\n「print」と入力することで前回の計算結果を表示することができます。\n「check」と入力することで(前回の計算結果)-1が素数かどうかをチェックします。")

input = input()

if input == "print":
    s = ''
    wb = openpyxl.load_workbook("test2.xlsx")
    ws = wb['Sheet']
    max = ws.max_row
    c = ws['C1'].value
    cell = ws.values
    list = list(cell)
    for i in range(max):
        if i != max-1 and i % 3 == 2:
            s = "," + str(list[i][0]) + s
        else:
            s = str(list[i][0]) + s
    print("2^"+str(c)+" =" )
    print(s)
    sys.exit()

if input == "check":
    start_time = time.perf_counter()

    s = ''
    wb = openpyxl.load_workbook("test2.xlsx")
    ws = wb['Sheet']
    a = int(ws['A1'].value)
    ws['A1'].value = a - 1
    max = ws.max_row
    c = ws['C1'].value
    cell = ws.values
    list = list(cell)
    for i in range(max):
        s = str(list[i][0]) + s
    print("2^"+str(c)+"-1 = "+str(s))

    l = len(s)

    if l % 2 == 0:
        o = s[:2]
        z = int(len(s[2:]))/2
    else:
        o = s[:3]
        z = int(len(s[3:]))/2

    p = math.sqrt(int(o))
    for i in range(0, int(z)):
        p = p * 10

    p = math.floor(p)

    if p % 2 == 0:
        p = p + 1

    # print("p1 = "+str(p))

    m = int(s) - p * p

    if m == 1:
        print(str(i)+"は素数です")
        sys.exit()

    q = p

    i = 1

    while True:
        o = str(p * q)
        if i < len(str(q)) or i < len(str(p)):
            if s[i] > o[i]:
                l = len(s) - i - len(str(q)) - 1
                if l != 0 and l > 0:
                    q = q + pow(10, l)
                    i = 0
            elif s[i] < o[i]:
                l = len(s) - i - len(str(p)) - 1
                if l != 0 and l > 0:
                    p = p - pow(10, l)
                    i = 0
            i = i + 1
        else:
            break
        # print("\r i = "+str(i)+" l = "+str(l)+" o = "+o, end="")

    # print("\n i = "+str(i)+" l = "+str(l)+" p = "+str(p)+" q = "+str(q)+" o = "+o)

    # 今日の収穫
    # xとyを一緒くたに動かすと、最後、xからyが出せる。
    # あとはこのxを動かす回数を求めてやればいい。直感ではただの2時方程式らしいんだけど、余計なものがついているのか立式が難しい。

    # q-p
    a = q - p
    # (q-p)^2-4*(s-p*q)
    b = (a**2) - 4 * (int(s)-(p*q))
    # 2p+2q
    c = 2 * p + 2 * q

    d = 2

    while True:
        d = d + 2
        if (d**2) + (c * d) + b > 0:
            e = math.sqrt((d**2) + (c * d) + b)
            if e.is_integer():
                # print("\r a = "+str(a)+" b = "+str(b)+" c = "+str(c)+" d = "+str(d), end="")
                break
        if p - d <= 2:
            print("\n"+str(s)+"は素数です")
            sys.exit()

    x = 0.5*(math.sqrt((d**2) + (c * d) + b)-d-a)

    p = p - int(x)
    q = q + int(x)

    while int(s) != p*q:
        if int(s) > p * q:
            q = q + 2
        elif int(s) < p * q:
            p = p - 2

        if p <= 2 or p == 0:
            print("\n"+str(s)+"は素数です")
            sys.exit()


    # print("\n p = "+str(p)+"q = "+str(q)+" p*q = "+str(p*q))

    if p * q == int(s):
        if p <= 2 or p == 0:
            print("\n"+str(s)+"は素数です")
        else:
            print("\n"+str(s)+"は、"+str(p)+"×"+str(q)+"です")
    else:
        print("\nなんかよくわからんがエラーだぞ")
        print("p = "+str(p)+" q = "+str(q)+" p*q = "+str(p*q)+" s = "+str(s))

    end_time = time.perf_counter()

    elapsed_time = end_time - start_time
    print("\n"+str(elapsed_time))

    sys.exit()

elif float(input).is_integer() == True:

    inputint = int(input)

    start_time = time.perf_counter()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'].value = 1
    ws['C1'].value = inputint

    max = ws.max_row

    for i in range(inputint):
        print("\r2^"+str(i+1)+" => Excel", end="")
        for j in range(max):
            a = ws['A'+str(j+1)].value
            if a is not None:
                r = a * 2
                if r > 9:
                    ws['A'+str(j+1)].value = int(str(r)[1])
                    ws['B'+str(j+2)].value = int(str(r)[0])
                else:
                    ws['A'+str(j+1)].value = int(r)

            max = ws.max_row

        for k in range(max):
            a = ws['A'+str(k+1)].value
            b = ws['B'+str(k+1)].value
            if a is not None and b is not None:
                ws['A'+str(k+1)].value = a + b
                ws['B'+str(k+1)].value = None
            elif a is None and b is not None:
                ws['A'+str(k+1)].value = b
                ws['B'+str(k+1)].value = None

    wb.save('test2.xlsx')

    end_time = time.perf_counter()

    elapsed_time = end_time - start_time
    print("\n"+str(elapsed_time))

else:
    print("error")
